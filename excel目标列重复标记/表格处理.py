# 左右列内容匹配，标黄处理
# A列：查询页面，B列：URL
# 目标表格名.xlsx：包含查询页面和URL的Excel文件
# 结果表格名.xlsx：处理后保存的Excel文件
# 运行前请确保安装 openpyxl 库：pip install openpyxl
# 运行前请确保关闭目标表格名.xlsx，否则会报错无法保存
# 运行后请检查结果表格名.xlsx，B列中包含A列查询页面的单元格已标黄
# 如果需要匹配任一查询页面（子串匹配）/ 完全匹配，请取消相应代码行的注释

import openpyxl  # 处理 Excel 文件  # openpyxl 只能处理 xlsx 格式
from openpyxl.styles import PatternFill # 用于单元格填充样式

# 打开 Excel 文件
file_path = "目标表格名.xlsx"
wb = openpyxl.load_workbook(file_path) # 加载工作簿
ws = wb.active # 选择活动工作表 
# 如果有多个工作表，可以用 wb['SheetName'] 选择特定工作表
# print(wb.sheetnames) # 打印所有工作表名称
# 黄底填充样式
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 第一步：提取 A列所有非空“查询页面”值，存入列表
query_pages = []
for row in ws.iter_rows(min_row=2):  # 跳过表头
    a_val = row[0].value  # A列值
    if a_val:
        query_pages.append(str(a_val).strip()) # 去除前后空格后存入列表

# 第二步：对每一行的 B列（URL）进行匹配，如果包含任一查询页面，则标黄
for row in ws.iter_rows(min_row=2):
    b_cell = row[1]
    b_val = str(b_cell.value).strip() if b_cell.value else ""

    # 判断是否匹配任一查询页面（子串匹配）
    # if any(query in b_val for query in query_pages):    # 匹配任一查询页面
    if b_val in query_pages: # 完全匹配
        b_cell.fill = yellow_fill # B列单元格标黄


# 保存文件
output_path = "结果表格名.xlsx"
wb.save(output_path)
print(f"✅ 已完成标黄处理并保存：{output_path}")
